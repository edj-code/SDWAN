"""
Unified vManage Policy Object Uploader & NGFW Policy Manager

Phase 1: Reads policy objects (Data Prefixes, Port Lists, Security Zones)
         from an Excel file and uploads them to vManage.
         - Security Zones support both Interface and VPN types.
Phase 2: Reads NGFW rules from the "NGFW Rules" sheet of the SAME Excel file
         and creates/updates the embedded-security profile and policy in vManage.
         
"""

import requests
import json
import re
import sys
import time
import os
import getpass
import urllib3
import openpyxl

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL CONFIGURATION (set after user input)
# ══════════════════════════════════════════════════════════════════════════════

VMANAGE_HOST = ""
BASE_URL = ""
USERNAME = ""
PASSWORD = ""
EXCEL_FILE = ""
NEW_POLICY_NAME = ""
NEW_POLICY_DESC = ""

# Mode: "create" or "update"
MODE = "create"

# For UPDATE mode (printed at the end of CREATE mode)
EXISTING_PROFILE_ID = ""
EXISTING_POLICY_ID = ""

# API tuning
API_TIMEOUT = 120
API_DELAY = 1
MAX_RETRIES = 3
RETRY_DELAY = 5
SKIP_PORT_OBJECTS = False

# Sheet name for NGFW rules in the Excel file
NGFW_RULES_SHEET = "NGFW Rules"

FALLBACK_DEFAULT_ACTION = "drop"

# ══════════════════════════════════════════════════════════════════════════════
# SHARED AUTHENTICATION
# ══════════════════════════════════════════════════════════════════════════════

def authenticate():
    """
    Authenticates to vManage and returns a requests.Session with
    JSESSIONID cookie and X-XSRF-TOKEN header set.
    """
    print("\n" + "=" * 70)
    print("[AUTH] Authenticating to vManage...")
    print("=" * 70)

    session = requests.Session()
    session.verify = False

    login_url = f"{BASE_URL}/j_security_check"
    resp = session.post(
        login_url,
        data={"j_username": USERNAME, "j_password": PASSWORD},
        timeout=30
    )
    if resp.status_code not in (200, 302, 303):
        print(f"[AUTH] ❌ Login failed. HTTP {resp.status_code}")
        sys.exit(1)

    cookies = resp.headers.get("Set-Cookie", "")
    if not cookies:
        print("[AUTH] ❌ Login failed. No session cookie returned. Check credentials.")
        sys.exit(1)

    print("[AUTH] ✅ JSESSIONID obtained.")

    token_url = f"{BASE_URL}/dataservice/client/token"
    resp = session.get(token_url, timeout=30)
    if resp.status_code != 200 or not resp.text.strip():
        print(f"[AUTH] ❌ Failed to get XSRF token. HTTP {resp.status_code}")
        sys.exit(1)

    token = resp.text.strip()
    session.headers.update({
        "X-XSRF-TOKEN": token,
        "Content-Type": "application/json",
        "Accept": "application/json"
    })

    print("[AUTH] ✅ XSRF token obtained.")
    print("[AUTH] ✅ Authentication successful.")
    return session


def re_auth(session):
    """Re-authenticate if the session expires mid-run."""
    session.post(
        f"{BASE_URL}/j_security_check",
        data={"j_username": USERNAME, "j_password": PASSWORD},
        timeout=30
    )
    resp = session.get(f"{BASE_URL}/dataservice/client/token", timeout=30)
    if resp.status_code == 200:
        token = resp.text.strip()
        session.headers.update({"X-XSRF-TOKEN": token})
        print("  [RE-AUTH] ✅ Re-authenticated")


# ══════════════════════════════════════════════════════════════════════════════
# SHARED API HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def api_get(session, url, description="API GET"):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, timeout=API_TIMEOUT)
            if resp.status_code == 200:
                return resp
            else:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return resp
        except Exception as e:
            print(f"  [{description}] ❌ Error: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return None
    return None


def api_post(session, url, payload, description="API POST"):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"  [{description}] Attempt {attempt}/{MAX_RETRIES}...")
            resp = session.post(url, json=payload, timeout=API_TIMEOUT)
            if resp.status_code in (200, 201):
                return resp
            elif resp.status_code == 429:
                wait = RETRY_DELAY * attempt
                print(f"  [{description}] ⚠️ Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            else:
                print(f"  [{description}] ❌ HTTP {resp.status_code}")
                print(f"  [{description}] Response: {resp.text[:500]}")
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return resp
        except requests.exceptions.Timeout:
            print(f"  [{description}] ⚠️ Timeout after {API_TIMEOUT}s")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return None
        except requests.exceptions.ConnectionError as e:
            print(f"  [{description}] ⚠️ Connection error: {e}")
            if attempt < MAX_RETRIES:
                try:
                    re_auth(session)
                except:
                    pass
                time.sleep(RETRY_DELAY)
                continue
            return None
        except Exception as e:
            print(f"  [{description}] ❌ Unexpected error: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return None
    return None


def api_put(session, url, payload, description="API PUT"):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"  [{description}] Attempt {attempt}/{MAX_RETRIES}...")
            resp = session.put(url, json=payload, timeout=API_TIMEOUT)
            if resp.status_code in (200, 201):
                return resp
            elif resp.status_code == 429:
                wait = RETRY_DELAY * attempt
                print(f"  [{description}] ⚠️ Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            else:
                print(f"  [{description}] ❌ HTTP {resp.status_code}")
                print(f"  [{description}] Response: {resp.text[:500]}")
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return resp
        except requests.exceptions.Timeout:
            print(f"  [{description}] ⚠️ Timeout")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return None
        except requests.exceptions.ConnectionError as e:
            print(f"  [{description}] ⚠️ Connection error: {e}")
            if attempt < MAX_RETRIES:
                try:
                    re_auth(session)
                except:
                    pass
                time.sleep(RETRY_DELAY)
                continue
            return None
        except Exception as e:
            print(f"  [{description}] ❌ Error: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return None
    return None


# ══════════════════════════════════════════════════════════════════════════════
# SHARED: GET POLICY-OBJECT FEATURE PROFILE ID
# ══════════════════════════════════════════════════════════════════════════════

def get_policy_object_feature_profile_id(session):
    """
    Finds the EXISTING global Feature Profile of type 'policy-object'.
    Used by both Phase 1 and Phase 2.
    """
    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/policy-object"
    print("\n[INIT] Searching for existing Policy Object Feature Profile...")

    try:
        response = session.get(url=url, timeout=API_TIMEOUT)
        if response.status_code == 200:
            data = response.json()
            profiles = []
            if isinstance(data, dict):
                profiles = data.get('data', [])
            elif isinstance(data, list):
                profiles = data

            if len(profiles) > 0:
                existing_profile = profiles[0]
                p_id = existing_profile.get('profileId')
                p_name = existing_profile.get('profileName')
                print(f"[INIT] ✅ Found existing profile '{p_name}'. ID: {p_id}")
                return p_id
            else:
                print("[INIT] ❌ No policy-object profiles found.")
                return None
        else:
            print(f"[INIT] ❌ API call failed. Status: {response.status_code}")
            print(f"       Response: {response.text[:500]}")
            return None
    except Exception as e:
        print(f"[INIT] ❌ Exception during profile search: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1: UPLOAD POLICY OBJECTS FROM EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def save_json_to_file(list_name, list_type, payload):
    """Save the JSON payload to a file for debugging/auditing."""
    output_dir = "generated_policy_objects"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    filename = os.path.join(output_dir, f"{list_type}_{list_name}.json")
    try:
        with open(filename, 'w') as f:
            json.dump(payload, f, indent=4)
        print(f"   [FILE] Saved JSON to {filename}")
    except IOError as e:
        print(f"   [ERROR] Could not save JSON file: {e}")


# ── Excel Reading (Phase 1 - Generic for Data Prefixes & Port Lists) ────────

def read_sheet_data(excel_file_path, sheet_name):
    """
    Reads a specific sheet from an Excel file (2-column format).
    Returns a list of tuples: (list_name, [values]).
    Used for Data Prefixes and Port Lists sheets.
    """
    data = []
    try:
        wb = openpyxl.load_workbook(excel_file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"   [WARNING] Sheet '{sheet_name}' not found in '{excel_file_path}'. Skipping.")
            wb.close()
            return data

        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or row[0] is None:
                continue
            list_name = str(row[0]).strip()
            list_name = "".join(c for c in list_name if c.isalnum() or c in ('_', '-'))
            raw_values = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            values = [v.strip() for v in raw_values.split(',') if v.strip()]
            if list_name and values:
                data.append((list_name, values))
        wb.close()
        print(f"   [READ] Found {len(data)} entries in sheet '{sheet_name}'.")
    except FileNotFoundError:
        print(f"   [ERROR] File '{excel_file_path}' not found.")
    except Exception as e:
        print(f"   [ERROR] Failed to read sheet '{sheet_name}': {e}")
    return data


# ── Excel Reading (Phase 1 - Security Zones with 3 columns) ─────────────────

def read_security_zones_data(excel_file_path):
    """
    Reads the 'Security Zones' sheet from an Excel file (3-column format).

    Expected layout:
        Col A = Zone Name
        Col B = Type ('VPN' or 'Interface')
        Col C = Comma-separated interfaces OR single VPN name

    Returns a list of tuples: (zone_name, zone_type, [values])
        - zone_type is normalized to 'vpn' or 'interface'
    """
    sheet_name = "Security Zones"
    data = []
    try:
        wb = openpyxl.load_workbook(excel_file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"   [WARNING] Sheet '{sheet_name}' not found in '{excel_file_path}'. Skipping.")
            wb.close()
            return data

        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or row[0] is None:
                continue

            # Column A: Zone Name
            zone_name = str(row[0]).strip()
            zone_name = "".join(c for c in zone_name if c.isalnum() or c in ('_', '-'))

            # Column B: Type (VPN or Interface)
            zone_type_raw = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ""

            # Column C: Values
            raw_values = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            values = [v.strip() for v in raw_values.split(',') if v.strip()]

            # Validate
            if not zone_name:
                continue

            if zone_type_raw not in ("vpn", "interface"):
                print(f"   [WARNING] Row {row_num}: Invalid type '{zone_type_raw}' for zone "
                      f"'{zone_name}'. Expected 'VPN' or 'Interface'. Skipping.")
                continue

            if not values:
                print(f"   [WARNING] Row {row_num}: No values for zone '{zone_name}'. Skipping.")
                continue

            data.append((zone_name, zone_type_raw, values))

        wb.close()
        print(f"   [READ] Found {len(data)} entries in sheet '{sheet_name}'.")

    except FileNotFoundError:
        print(f"   [ERROR] File '{excel_file_path}' not found.")
    except Exception as e:
        print(f"   [ERROR] Failed to read sheet '{sheet_name}': {e}")
    return data


# ── Parcel Creation (Phase 1) ────────────────────────────────────────────────

def create_policy_object_parcel(session, parent_id, list_type, list_name, payload):
    """POST a policy-object parcel to vManage."""
    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/policy-object/{parent_id}/{list_type}"
    save_json_to_file(list_name, list_type, payload)

    try:
        print(f"   [API] Sending POST request for '{list_name}' (type: {list_type})...")
        response = session.post(url=url, json=payload, timeout=API_TIMEOUT)
        if response.status_code in [200, 201]:
            resp_json = response.json()
            obj_id = resp_json.get('id')
            print(f"   [SUCCESS] Created '{list_name}' (type: {list_type}). Parcel ID: {obj_id}")
        elif response.status_code == 400:
            print(f"   [FAIL] Error 400: Schema validation failed for '{list_name}'.")
            print(f"   Response: {response.text}")
        else:
            print(f"   [FAIL] Error {response.status_code} for '{list_name}'.")
            print(f"   Response: {response.text}")
    except requests.exceptions.RequestException as e:
        print(f"   [ERROR] Request error for '{list_name}': {e}")


# ── Payload Builders (Phase 1) ───────────────────────────────────────────────

def build_security_data_ip_prefix_payload(list_name, values):
    formatted_entries = []
    for prefix in values:
        entry = {"ipPrefix": {"optionType": "global", "value": prefix}}
        formatted_entries.append(entry)
    return {
        "name": list_name,
        "description": f"Security Data IP Prefix List: {list_name}",
        "data": {"entries": formatted_entries}
    }


def build_security_port_payload(list_name, values):
    formatted_entries = []
    for port in values:
        entry = {"port": {"optionType": "global", "value": port}}
        formatted_entries.append(entry)
    return {
        "name": list_name,
        "description": f"Security Port List: {list_name}",
        "data": {"entries": formatted_entries}
    }


def build_security_zone_payload(zone_name, zone_type, values):
    """
    Builds the JSON payload for 'security-zone' lists.

    Args:
        zone_name: Name of the security zone
        zone_type: 'interface' or 'vpn'
        values:    List of interface names (for interface type)
                   OR list with a single VPN name (for vpn type)

    Interface type payload example:
        {
            "name": "INSIDE_ZONE",
            "description": "Security Zone: INSIDE_ZONE",
            "data": {
                "entries": [
                    {"interface": {"optionType": "global", "value": "GigabitEthernet0/0/0"}},
                    {"interface": {"optionType": "global", "value": "GigabitEthernet0/0/1"}}
                ]
            }
        }

    VPN type payload example:
        {
            "name": "VPN10_ZONE",
            "description": "Security Zone: VPN10_ZONE",
            "data": {
                "entries": [
                    {"vpn": {"optionType": "global", "value": "VPN10"}}
                ]
            }
        }
    """
    formatted_entries = []

    if zone_type == "interface":
        for interface_name in values:
            entry = {
                "interface": {
                    "optionType": "global",
                    "value": interface_name
                }
            }
            formatted_entries.append(entry)
    elif zone_type == "vpn":
        # VPN type: typically a single VPN name
        vpn_name = values[0]
        entry = {
            "vpn": {
                "optionType": "global",
                "value": vpn_name
            }
        }
        formatted_entries.append(entry)

    payload = {
        "name": zone_name,
        "description": f"Security Zone: {zone_name}",
        "data": {
            "entries": formatted_entries
        }
    }
    return payload


# ── Sheet Processing (Phase 1) ───────────────────────────────────────────────

def process_sheet(session, parent_id, excel_file, sheet_name, list_type, payload_builder):
    """
    Read a 2-column sheet and create policy object parcels for each row.
    Used for Data Prefixes and Port Lists.
    """
    print(f"\n{'=' * 60}")
    print(f"Processing sheet: '{sheet_name}' -> list type: '{list_type}'")
    print(f"{'=' * 60}")

    entries = read_sheet_data(excel_file, sheet_name)
    if not entries:
        print(f"   [INFO] No entries to process for sheet '{sheet_name}'.")
        return

    for idx, (list_name, values) in enumerate(entries, start=1):
        print(f"\n--- [{sheet_name}] Row {idx}: '{list_name}' with {len(values)} value(s) ---")
        payload = payload_builder(list_name, values)
        create_policy_object_parcel(session, parent_id, list_type, list_name, payload)


def process_security_zones_sheet(session, parent_id, excel_file):
    """
    Read the 3-column 'Security Zones' sheet and create security-zone
    parcels for each row, supporting both Interface and VPN types.
    """
    list_type = "security-zone"

    print(f"\n{'=' * 60}")
    print(f"Processing sheet: 'Security Zones' -> list type: '{list_type}'")
    print(f"  Supports: Interface entries and VPN entries")
    print(f"{'=' * 60}")

    entries = read_security_zones_data(excel_file)
    if not entries:
        print(f"   [INFO] No entries to process for sheet 'Security Zones'.")
        return

    for idx, (zone_name, zone_type, values) in enumerate(entries, start=1):
        type_label = zone_type.upper()
        if zone_type == "interface":
            print(f"\n--- [Security Zones] Row {idx}: '{zone_name}' "
                  f"[{type_label}] with {len(values)} interface(s) ---")
            for v in values:
                print(f"      • {v}")
        else:
            print(f"\n--- [Security Zones] Row {idx}: '{zone_name}' "
                  f"[{type_label}] -> VPN: {values[0]} ---")

        payload = build_security_zone_payload(zone_name, zone_type, values)
        create_policy_object_parcel(session, parent_id, list_type, zone_name, payload)


def run_phase1(session, parent_profile_id):
    """
    Phase 1: Upload policy objects from Excel.
    """
    print("\n")
    print("#" * 70)
    print("#  PHASE 1: Upload Policy Objects from Excel")
    print(f"#  Excel File: {EXCEL_FILE}")
    print("#" * 70)

    # Process Data Prefixes (2-column format)
    process_sheet(
        session=session,
        parent_id=parent_profile_id,
        excel_file=EXCEL_FILE,
        sheet_name="Data Prefixes",
        list_type="security-data-ip-prefix",
        payload_builder=build_security_data_ip_prefix_payload
    )

    # Process Port Lists (2-column format)
    process_sheet(
        session=session,
        parent_id=parent_profile_id,
        excel_file=EXCEL_FILE,
        sheet_name="Port Lists",
        list_type="security-port",
        payload_builder=build_security_port_payload
    )

    # Process Security Zones (3-column format with VPN/Interface type)
    process_security_zones_sheet(
        session=session,
        parent_id=parent_profile_id,
        excel_file=EXCEL_FILE
    )

    print(f"\n{'=' * 60}")
    print("[PHASE 1] ✅ All sheets processed successfully.")
    print(f"{'=' * 60}")


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2: CREATE/UPDATE NGFW POLICIES FROM EXCEL SHEET "NGFW Rules"
# ══════════════════════════════════════════════════════════════════════════════

# ── Validation Helpers ────────────────────────────────────────────────────────

def validate_cidr(cidr: str) -> bool:
    try:
        import ipaddress
        ipaddress.ip_network(cidr, strict=False)
        return True
    except (ValueError, AttributeError):
        return False


def ensure_cidr(ip_str: str) -> str:
    ip_str = ip_str.strip()
    if "/" not in ip_str:
        ip_str = ip_str + "/32"
    return ip_str


def is_empty_or_dash(value: str) -> bool:
    return not value or value.strip() == "" or value.strip() == "-"


def sanitize_ip_list(raw: str) -> list:
    if is_empty_or_dash(raw):
        return []
    normalized = raw.replace(";", ",")
    parts = [p.strip() for p in normalized.split(",") if p.strip()]
    clean = []
    for p in parts:
        cidr = ensure_cidr(p)
        if validate_cidr(cidr):
            clean.append(cidr)
    return clean


def sanitize_port_value(raw: str) -> list:
    if is_empty_or_dash(raw):
        return []
    normalized = raw.replace(";", ",")
    parts = [p.strip() for p in normalized.split(",") if p.strip()]
    clean = []
    for p in parts:
        if re.match(r'^\d+(-\d+)?$', p):
            if '-' in p:
                start, end = p.split('-')
                if 1 <= int(start) <= 65535 and 1 <= int(end) <= 65535:
                    clean.append(p)
            else:
                if 1 <= int(p) <= 65535:
                    clean.append(p)
    return clean


# ── Field Processors ─────────────────────────────────────────────────────────

def process_ip_field(field_type, field_data, cache, field_name):
    if is_empty_or_dash(field_data):
        return None
    if field_type == "object":
        object_names = [n.strip() for n in re.split(r'[;,]', field_data) if n.strip()]
        uuid_list = []
        for obj_name in object_names:
            uuid = cache.resolve("security-data-ip-prefix", obj_name)
            if not uuid:
                return None
            uuid_list.append(uuid)
        if field_name == "source":
            return {"sourceDataPrefixList": {"refId": {"optionType": "global", "value": uuid_list}}}
        else:
            return {"destinationDataPrefixList": {"refId": {"optionType": "global", "value": uuid_list}}}
    elif field_type == "value":
        ip_list = sanitize_ip_list(field_data)
        if not ip_list:
            return None
        if field_name == "source":
            return {"sourceIp": {"ipv4Value": {"optionType": "global", "value": ip_list}}}
        else:
            return {"destinationIp": {"ipv4Value": {"optionType": "global", "value": ip_list}}}
    return None


def process_port_field(port_type, port_data, cache, field_name):
    if is_empty_or_dash(port_data):
        return None
    if port_type == "object":
        if SKIP_PORT_OBJECTS:
            return None
        port_names = [n.strip() for n in re.split(r'[;,]', port_data) if n.strip()]
        uuid_list = []
        for pname in port_names:
            uuid = cache.resolve("security-port", pname)
            if not uuid:
                print(f"  ⚠️ Port object not found: '{pname}'")
                return None
            uuid_list.append(uuid)
        if field_name == "source":
            return {"sourcePortList": {"refId": {"value": uuid_list, "optionType": "global"}}}
        else:
            return {"destinationPortList": {"refId": {"value": uuid_list, "optionType": "global"}}}
    elif port_type == "value":
        ports = sanitize_port_value(port_data)
        if not ports:
            return None
        if field_name == "source":
            return {"sourcePort": {"portValue": {"optionType": "global", "value": ports}}}
        else:
            return {"destinationPort": {"portValue": {"optionType": "global", "value": ports}}}
    return None


def process_protocol_field(proto_type, proto_data, cache):
    if is_empty_or_dash(proto_data):
        return None
    if proto_type == "object":
        proto_names = [n.strip() for n in re.split(r'[;,]', proto_data) if n.strip()]
        uuid_list = []
        for pname in proto_names:
            uuid = cache.resolve("security-protocolname", pname)
            if not uuid:
                return None
            uuid_list.append(uuid)
        return {"protocolNameList": {"refId": {"optionType": "global", "value": uuid_list}}}
    elif proto_type in ("value", "name"):
        proto_names = [p.strip().lower() for p in re.split(r'[;,]', proto_data) if p.strip()]
        if proto_names:
            return {"protocolName": {"optionType": "global", "value": proto_names}}
    return None


# ── Object Cache ──────────────────────────────────────────────────────────────

class ObjectCache:
    LIST_TYPES = [
        "security-data-ip-prefix",
        "security-port",
        "security-zone",
        "security-protocolname"
    ]

    def __init__(self, session, policy_object_id):
        self.session = session
        self.policy_object_id = policy_object_id
        self.cache = {}
        self._load_all()

    def _load_all(self):
        print("\n[CACHE] Loading policy objects into cache...")
        for list_type in self.LIST_TYPES:
            url = (
                f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/policy-object"
                f"/{self.policy_object_id}/{list_type}"
            )
            try:
                resp = self.session.get(url, timeout=API_TIMEOUT)
                if resp.status_code != 200:
                    continue
                items = resp.json()
                if isinstance(items, dict):
                    items = items.get("data", [])
                count = 0
                for item in items:
                    pid = item.get("parcelId")
                    name = item.get("name") or item.get("payload", {}).get("name")
                    if name and pid:
                        self.cache[(list_type, name.lower())] = pid
                        count += 1
                print(f"[CACHE] {list_type}: {count}")
            except Exception as e:
                print(f"[CACHE] ❌ Error loading {list_type}: {e}")
        print(f"[CACHE] ✅ {len(self.cache)} total objects loaded")

    def resolve(self, list_type, name):
        return self.cache.get((list_type, name.strip().lower()))


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL PARSING FOR NGFW RULES
# ══════════════════════════════════════════════════════════════════════════════

def cell_to_str(cell_value):
    """Safely convert an Excel cell value to a stripped string."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def parse_ngfw_rules_from_excel(excel_file_path, cache):
    """
    Reads the 'NGFW Rules' sheet from the Excel file and returns:
        zone_pair_rules:    dict of (src_zone, dst_zone) -> [sequence_dicts]
        zone_pair_defaults: dict of (src_zone, dst_zone) -> default action string

    Column layout (Row 1 = header, skipped):
        Col A  (0)  = Source Zone
        Col B  (1)  = Destination Zone
        Col C  (2)  = Source IP Type
        Col D  (3)  = Source IP Data
        Col E  (4)  = Source Port Type
        Col F  (5)  = Source Port Data
        Col G  (6)  = Destination IP Type
        Col H  (7)  = Destination IP Data
        Col I  (8)  = Destination Port Type
        Col J  (9)  = Destination Port Data
        Col K  (10) = Protocol Type
        Col L  (11) = Protocol Data
        Col M  (12) = Action (inspect / pass / drop)
        Col N  (13) = Default Action for zone pair (drop / pass / inspect) [optional]
    """
    print(f"\n[EXCEL] Parsing NGFW rules from sheet '{NGFW_RULES_SHEET}'...")

    zone_pair_rules = {}
    zone_pair_defaults = {}  # NEW: stores default action per zone pair
    any_any_converted = 0
    skipped_rows = 0
    total_data_rows = 0

    try:
        wb = openpyxl.load_workbook(excel_file_path, read_only=True)

        if NGFW_RULES_SHEET not in wb.sheetnames:
            print(f"[EXCEL] ❌ Sheet '{NGFW_RULES_SHEET}' not found in '{excel_file_path}'.")
            print(f"[EXCEL]    Available sheets: {wb.sheetnames}")
            wb.close()
            return {}, {}

        ws = wb[NGFW_RULES_SHEET]

        for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Skip header row
            if row_num == 1:
                continue

            # Skip completely empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Pad the row to at least 14 columns (A through N)
            row_data = [cell_to_str(cell) for cell in row]
            while len(row_data) < 14:
                row_data.append("")

            total_data_rows += 1

            src_zone = row_data[0]
            dst_zone = row_data[1]
            action = row_data[12].lower()
            default_action_raw = row_data[13].lower()  # NEW: Column N

            if action not in ("inspect", "pass", "drop"):
                skipped_rows += 1
                continue

            if is_empty_or_dash(src_zone) or is_empty_or_dash(dst_zone):
                skipped_rows += 1
                continue

            zone_pair = (src_zone, dst_zone)

            # NEW: Capture default action from the FIRST row that specifies it
            if zone_pair not in zone_pair_defaults:
                if default_action_raw in ("drop", "pass", "inspect"):
                    zone_pair_defaults[zone_pair] = default_action_raw
                    print(f"  [EXCEL] ℹ️ Default action for {src_zone} -> {dst_zone}: "
                          f"{default_action_raw.upper()}")
                else:
                    # No default specified yet — will use fallback later
                    pass
            else:
                # Zone pair already has a default; if this row also specifies one, warn
                if default_action_raw in ("drop", "pass", "inspect"):
                    if default_action_raw != zone_pair_defaults[zone_pair]:
                        print(f"  [EXCEL] ⚠️ Row {row_num}: Conflicting default action "
                              f"'{default_action_raw}' for {src_zone} -> {dst_zone} "
                              f"(using first: '{zone_pair_defaults[zone_pair]}')")

            entries = []
            failed = False

            # Source IP (columns C=2, D=3)
            if not is_empty_or_dash(row_data[3]):
                e = process_ip_field(row_data[2].lower(), row_data[3], cache, "source")
                if e:
                    entries.append(e)
                else:
                    failed = True

            # Source Port (columns E=4, F=5)
            if not is_empty_or_dash(row_data[5]) and not failed:
                if not (row_data[4].lower() == "object" and SKIP_PORT_OBJECTS):
                    e = process_port_field(row_data[4].lower(), row_data[5], cache, "source")
                    if e:
                        entries.append(e)

            # Destination IP (columns G=6, H=7)
            if not is_empty_or_dash(row_data[7]) and not failed:
                e = process_ip_field(row_data[6].lower(), row_data[7], cache, "destination")
                if e:
                    entries.append(e)
                else:
                    failed = True

            # Destination Port (columns I=8, J=9)
            if not is_empty_or_dash(row_data[9]) and not failed:
                if not (row_data[8].lower() == "object" and SKIP_PORT_OBJECTS):
                    e = process_port_field(row_data[8].lower(), row_data[9], cache, "destination")
                    if e:
                        entries.append(e)

            # Protocol (columns K=10, L=11)
            if not is_empty_or_dash(row_data[11]) and not failed:
                e = process_protocol_field(row_data[10].lower(), row_data[11], cache)
                if e:
                    entries.append(e)
                else:
                    failed = True

            if failed:
                skipped_rows += 1
                print(f"  [EXCEL] ⚠️ Row {row_num}: Skipped (unresolved object)")
                continue

            if not entries:
                any_any_converted += 1
                entries = []

            if zone_pair not in zone_pair_rules:
                zone_pair_rules[zone_pair] = []

            rule_number = len(zone_pair_rules[zone_pair]) + 1
            seq = {
                "actions": [],
                "baseAction": {"optionType": "global", "value": action},
                "disableSequence": {"optionType": "global", "value": False},
                "match": {"entries": entries},
                "sequenceId": {"optionType": "global", "value": str(rule_number)},
                "sequenceName": {"optionType": "global", "value": f"Rule{rule_number}"},
                "sequenceType": {"optionType": "global", "value": "ngfirewall"}
            }
            zone_pair_rules[zone_pair].append(seq)

        wb.close()

    except FileNotFoundError:
        print(f"[EXCEL] ❌ File '{excel_file_path}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"[EXCEL] ❌ Error reading Excel: {e}")
        sys.exit(1)

    # Apply fallback default for any zone pair that didn't specify one
    for zp in zone_pair_rules:
        if zp not in zone_pair_defaults:
            zone_pair_defaults[zp] = "drop"  # Global fallback
            print(f"  [EXCEL] ℹ️ No default action for {zp[0]} -> {zp[1]}, "
                  f"using fallback: DROP")

    total_rules = sum(len(r) for r in zone_pair_rules.values())
    print(f"[EXCEL] ✅ Read {total_data_rows} data rows from '{NGFW_RULES_SHEET}'")
    print(f"[EXCEL] ✅ {len(zone_pair_rules)} zone pairs, {total_rules} total rules")
    if any_any_converted > 0:
        print(f"[EXCEL] ℹ️ {any_any_converted} any/any rules (empty match entries)")
    if skipped_rows > 0:
        print(f"[EXCEL] ⚠️ {skipped_rows} rows skipped")

    print(f"\n[EXCEL] Zone pair summary:")
    for zp, rules in zone_pair_rules.items():
        default = zone_pair_defaults.get(zp, "drop")
        print(f"  {zp[0]} -> {zp[1]}: {len(rules)} rules | default: {default.upper()}")

    return zone_pair_rules, zone_pair_defaults


# ── vManage Operations (Phase 2) ─────────────────────────────────────────────

def create_new_security_profile(session):
    print("\n[CREATE PROFILE] Creating embedded-security profile...")
    payload = {"name": NEW_POLICY_NAME, "description": NEW_POLICY_DESC}
    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security"
    resp = api_post(session, url, payload, "CREATE PROFILE")
    if resp and resp.status_code in (200, 201):
        result = resp.json()
        profile_id = result.get("profileId") or result.get("id")
        print(f"[CREATE PROFILE] ✅ {profile_id}")
        return profile_id
    print("[CREATE PROFILE] ❌ Failed")
    return None


def get_existing_ngfw_parcels(session, security_profile_id):
    """Fetch existing NGFW parcels: name -> parcelId."""
    print(f"\n[EXISTING NGFW] Fetching existing parcels...")
    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security/{security_profile_id}/unified/ngfirewall"
    resp = api_get(session, url, "GET EXISTING NGFW")
    existing = {}
    if resp and resp.status_code == 200:
        data = resp.json()
        parcels = data.get("data", []) if isinstance(data, dict) else data
        if isinstance(parcels, list):
            for p in parcels:
                pid = p.get("parcelId")
                name = p.get("payload", {}).get("name", "")
                if name and pid:
                    existing[name] = pid
        print(f"[EXISTING NGFW] ✅ Found {len(existing)} existing parcels")
        for name, pid in existing.items():
            print(f"  ♻️ {name} -> {pid}")
    return existing


def get_existing_policy(session, security_profile_id):
    """Fetch existing policy assembly."""
    print(f"\n[EXISTING POLICY] Fetching existing policy...")
    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security/{security_profile_id}/policy"
    resp = api_get(session, url, "GET EXISTING POLICY")
    if resp and resp.status_code == 200:
        data = resp.json()
        items = data.get("data", []) if isinstance(data, dict) else data
        if isinstance(items, list) and len(items) > 0:
            policy = items[0]
            policy_id = policy.get("parcelId")
            assembly = policy.get("payload", {}).get("data", {}).get("assembly", [])
            print(f"[EXISTING POLICY] ✅ Policy ID: {policy_id}")
            print(f"[EXISTING POLICY] ✅ Existing assembly entries: {len(assembly)}")
            return policy_id, assembly
    print(f"[EXISTING POLICY] ⚠️ No existing policy found")
    return None, []


def create_ngfw_parcel(session, security_profile_id, zone_pair, sequences, 
                       index, total, default_action="drop"):
    """
    Creates an NGFW parcel with a configurable default action.
    
    Args:
        default_action: 'drop', 'pass', or 'inspect' — applied when no rule matches.
    """
    src_zone, dst_zone = zone_pair
    parcel_name = f"NGFW_{src_zone}_to_{dst_zone}"
    print(f"\n[CREATE NGFW {index}/{total}] {parcel_name} "
          f"({len(sequences)} rules, default: {default_action.upper()})")

    payload = {
        "name": parcel_name,
        "description": f"Rules: {src_zone} to {dst_zone}",
        "optimized": False,
        "containsUtd": False,
        "containsTls": False,
        "data": {
            "defaultActionType": {
                "optionType": "global", 
                "value": default_action  # NOW CONFIGURABLE
            },
            "sequences": sequences
        }
    }

    url = (f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security"
           f"/{security_profile_id}/unified/ngfirewall")
    resp = api_post(session, url, payload, f"NGFW {index}/{total}")

    if resp and resp.status_code in (200, 201):
        result = resp.json()
        parcel_id = result.get("parcelId")
        print(f"[CREATE NGFW {index}/{total}] ✅ {parcel_id}")
        return parcel_id
    else:
        print(f"[CREATE NGFW {index}/{total}] ❌ Failed")
        if resp:
            print(f"  Response: {resp.text[:500]}")
        return None


def build_assembly_entry(src_zone_name, dst_zone_name, ngfw_id, cache):
    """Build a single assembly entry."""
    src_zone_uuid = cache.resolve("security-zone", src_zone_name)
    if not src_zone_uuid:
        print(f"  ⚠️ Zone '{src_zone_name}' not found")
        return None

    if dst_zone_name.lower() == "self":
        dst_zone_value = {"optionType": "global", "value": "self"}
    else:
        dst_zone_uuid = cache.resolve("security-zone", dst_zone_name)
        if not dst_zone_uuid:
            print(f"  ⚠️ Zone '{dst_zone_name}' not found")
            return None
        dst_zone_value = {"refId": {"optionType": "global", "value": dst_zone_uuid}}

    return {
        "ngfirewall": {
            "refId": {"optionType": "global", "value": ngfw_id},
            "entries": [{
                "srcZone": {"refId": {"optionType": "global", "value": src_zone_uuid}},
                "dstZone": dst_zone_value
            }]
        }
    }


def create_policy(session, security_profile_id, assembly):
    """Create a NEW policy."""
    payload = {
        "name": NEW_POLICY_NAME + "_Policy",
        "description": "Security policy",
        "data": {
            "settings": {
                "securityLogging": {"optionType": "network-settings", "value": True}
            },
            "assembly": assembly
        }
    }

    with open("debug_policy_payload.json", "w") as f:
        json.dump(payload, f, indent=2)
    print(f"[CREATE POLICY] 💾 Saved to: debug_policy_payload.json")
    print(f"[CREATE POLICY] Assembly entries: {len(assembly)}")

    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security/{security_profile_id}/policy"
    resp = api_post(session, url, payload, "CREATE POLICY")

    if resp and resp.status_code in (200, 201):
        result = resp.json()
        policy_id = result.get("parcelId") or result.get("id")
        print(f"[CREATE POLICY] ✅ Policy ID: {policy_id}")
        return policy_id
    else:
        print(f"[CREATE POLICY] ❌ Failed")
        if resp:
            print(f"  Response: {resp.text[:1000]}")
        return None


def update_policy(session, security_profile_id, policy_id, assembly):
    """Update an EXISTING policy."""
    payload = {
        "name": NEW_POLICY_NAME + "_Policy",
        "description": "Security policy",
        "data": {
            "settings": {
                "securityLogging": {"optionType": "network-settings", "value": True}
            },
            "assembly": assembly
        }
    }

    with open("debug_policy_update.json", "w") as f:
        json.dump(payload, f, indent=2)
    print(f"[UPDATE POLICY] 💾 Saved to: debug_policy_update.json")
    print(f"[UPDATE POLICY] Assembly entries: {len(assembly)}")

    url = f"{BASE_URL}/dataservice/v1/feature-profile/sdwan/embedded-security/{security_profile_id}/policy/{policy_id}"
    resp = api_put(session, url, payload, "UPDATE POLICY")

    if resp and resp.status_code in (200, 201):
        print(f"[UPDATE POLICY] ✅ Success")
        return True
    else:
        print(f"[UPDATE POLICY] ❌ Failed")
        if resp:
            print(f"  Response: {resp.text[:1000]}")
        return False


# ── CREATE Mode (Phase 2) ────────────────────────────────────────────────────

def run_phase2_create(session, cache, zone_pair_rules, zone_pair_defaults):
    """Create new profile + parcels + policy from scratch."""
    start_time = time.time()

    print(f"\n{'=' * 70}")
    print(f"[PHASE 2 - CREATE] Creating everything from scratch")
    print(f"{'=' * 70}")

    security_profile_id = create_new_security_profile(session)
    if not security_profile_id:
        print("[PHASE 2] ❌ Failed to create security profile. Exiting Phase 2.")
        return

    total_pairs = len(zone_pair_rules)
    ngfw_parcels = []
    failed_parcels = []

    for idx, (zone_pair, sequences) in enumerate(zone_pair_rules.items(), 1):
        # Look up the default action for THIS specific zone pair
        default_action = zone_pair_defaults.get(zone_pair, "drop")

        ngfw_id = create_ngfw_parcel(
            session, security_profile_id, zone_pair, sequences,
            idx, total_pairs,
            default_action=default_action    # ← THIS WAS MISSING
        )
        if ngfw_id:
            ngfw_parcels.append((zone_pair, ngfw_id))
        else:
            failed_parcels.append(zone_pair)

        elapsed = time.time() - start_time
        pct = (idx / total_pairs) * 100
        print(f"  📊 Progress: {idx}/{total_pairs} ({pct:.0f}%) | "
              f"✅ {len(ngfw_parcels)} | ❌ {len(failed_parcels)} | ⏱ {elapsed:.0f}s")
        if idx < total_pairs:
            time.sleep(API_DELAY)

    if not ngfw_parcels:
        print("[PHASE 2] ❌ No parcels created. Exiting Phase 2.")
        return

    assembly = []
    for (src_zone, dst_zone), ngfw_id in ngfw_parcels:
        entry = build_assembly_entry(src_zone, dst_zone, ngfw_id, cache)
        if entry:
            assembly.append(entry)

    policy_id = create_policy(session, security_profile_id, assembly)

    elapsed = time.time() - start_time

    print(f"\n{'#' * 70}")
    if policy_id:
        print(f"#  ✅ PHASE 2 CREATE COMPLETE!")
    else:
        print(f"#  ⚠️ PARCELS CREATED BUT POLICY FAILED")
    print(f"#")
    print(f"#  Profile ID: {security_profile_id}")
    print(f"#  Policy ID:  {policy_id}")
    print(f"#  Parcels:    {len(ngfw_parcels)}/{total_pairs}")
    print(f"#  Time:       {elapsed:.0f}s ({elapsed / 60:.1f}m)")
    print(f"#")
    print(f"#  Default Actions Used:")
    for (src, dst), _ in ngfw_parcels:
        da = zone_pair_defaults.get((src, dst), "drop")
        print(f"#    {src} -> {dst}: {da.upper()}")
    print(f"#")
    print(f"#  ─── FOR FUTURE UPDATES ───────────────────────────")
    print(f"#  While running again the script, set these values when asked to, and change MODE to 'update':")
    print(f"#")
    print(f"#  MODE = \"update\"")
    print(f"#  EXISTING_PROFILE_ID = \"{security_profile_id}\"")
    print(f"#  EXISTING_POLICY_ID  = \"{policy_id}\"")
    print(f"#")
    print(f"{'#' * 70}")

    if failed_parcels:
        print(f"\n  ❌ Failed parcels:")
        for fp in failed_parcels:
            print(f"    - {fp[0]} -> {fp[1]}")


# ── UPDATE Mode (Phase 2) ────────────────────────────────────────────────────

def run_phase2_update(session, cache, zone_pair_rules, zone_pair_defaults):
    """Add new parcels from Excel to existing profile, update policy."""
    start_time = time.time()

    if not EXISTING_PROFILE_ID:
        print("[UPDATE] ❌ EXISTING_PROFILE_ID is not set!")
        print("[UPDATE] Run in CREATE mode first, then copy the IDs.")
        return

    security_profile_id = EXISTING_PROFILE_ID

    print(f"\n{'=' * 70}")
    print(f"[PHASE 2 - UPDATE] Adding new rules to existing profile")
    print(f"  Profile: {security_profile_id}")
    print(f"{'=' * 70}")

    existing_ngfw = get_existing_ngfw_parcels(session, security_profile_id)

    policy_id, existing_assembly = get_existing_policy(session, security_profile_id)

    if not policy_id and EXISTING_POLICY_ID:
        policy_id = EXISTING_POLICY_ID
        print(f"[UPDATE] Using provided policy ID: {policy_id}")

    if not policy_id:
        print("[UPDATE] ❌ No existing policy found and EXISTING_POLICY_ID not set!")
        return

    total_pairs = len(zone_pair_rules)
    new_parcels = []
    reused_parcels = []
    failed_parcels = []

    print(f"\n[UPDATE] Processing {total_pairs} zone pairs from Excel...")

    for idx, (zone_pair, sequences) in enumerate(zone_pair_rules.items(), 1):
        src_zone, dst_zone = zone_pair
        parcel_name = f"NGFW_{src_zone}_to_{dst_zone}"

        if parcel_name in existing_ngfw:
            existing_id = existing_ngfw[parcel_name]
            print(f"\n[NGFW {idx}/{total_pairs}] ♻️ EXISTS: {parcel_name} -> {existing_id}")
            reused_parcels.append((zone_pair, existing_id))
        else:
            # Look up the default action for THIS specific zone pair
            default_action = zone_pair_defaults.get(zone_pair, "drop")

            ngfw_id = create_ngfw_parcel(
                session, security_profile_id, zone_pair, sequences,
                idx, total_pairs,
                default_action=default_action    # ← THIS WAS MISSING
            )
            if ngfw_id:
                new_parcels.append((zone_pair, ngfw_id))
            else:
                failed_parcels.append(zone_pair)
            time.sleep(API_DELAY)

        elapsed = time.time() - start_time
        pct = (idx / total_pairs) * 100
        print(f"  📊 Progress: {idx}/{total_pairs} ({pct:.0f}%) | "
              f"♻️ {len(reused_parcels)} reused | ✅ {len(new_parcels)} new | "
              f"❌ {len(failed_parcels)} failed | ⏱ {elapsed:.0f}s")

    print(f"\n[UPDATE] Parcel Summary:")
    print(f"  ♻️ Reused:  {len(reused_parcels)}")
    print(f"  ✅ Created: {len(new_parcels)}")
    print(f"  ❌ Failed:  {len(failed_parcels)}")

    all_parcels = reused_parcels + new_parcels

    if not all_parcels:
        print("[UPDATE] ❌ No parcels available. Exiting.")
        return

    print(f"\n[UPDATE] Building full assembly ({len(all_parcels)} zone pairs)...")

    full_assembly = []
    for (src_zone, dst_zone), ngfw_id in all_parcels:
        entry = build_assembly_entry(src_zone, dst_zone, ngfw_id, cache)
        if entry:
            full_assembly.append(entry)
            da = zone_pair_defaults.get((src_zone, dst_zone), "drop")
            print(f"  ✅ {src_zone} -> {dst_zone} (default: {da.upper()})")

    if not full_assembly:
        print("[UPDATE] ❌ No valid assembly entries. Exiting.")
        return

    print(f"\n[UPDATE] Updating policy {policy_id} with {len(full_assembly)} entries...")
    success = update_policy(session, security_profile_id, policy_id, full_assembly)

    elapsed = time.time() - start_time

    print(f"\n{'#' * 70}")
    if success:
        print(f"#  ✅ UPDATE COMPLETE!")
    else:
        print(f"#  ❌ UPDATE FAILED")
    print(f"#")
    print(f"#  Profile ID: {security_profile_id}")
    print(f"#  Policy ID:  {policy_id}")
    print(f"#  Reused:     {len(reused_parcels)} parcels")
    print(f"#  New:        {len(new_parcels)} parcels")
    print(f"#  Total:      {len(full_assembly)} assembly entries")
    print(f"#  Time:       {elapsed:.0f}s ({elapsed / 60:.1f}m)")
    print(f"#")
    print(f"#  Default Actions Used:")
    for (src, dst), _ in all_parcels:
        da = zone_pair_defaults.get((src, dst), "drop")
        print(f"#    {src} -> {dst}: {da.upper()}")
    print(f"#")
    print(f"#  NEXT: Push the config group to the device in vManage")
    print(f"{'#' * 70}")

    if failed_parcels:
        print(f"\n  ❌ Failed parcels:")
        for fp in failed_parcels:
            print(f"    - {fp[0]} -> {fp[1]}")


def run_phase2(session, policy_object_profile_id):
    """
    Phase 2: Create/Update NGFW policies from the 'NGFW Rules' sheet
    of the same Excel file used in Phase 1.
    """
    print("\n")
    print("#" * 70)
    print("#  PHASE 2: Create/Update NGFW Policies")
    print(f"#  Excel File: {EXCEL_FILE}")
    print(f"#  Sheet:      {NGFW_RULES_SHEET}")
    print(f"#  Mode:       {MODE.upper()}")
    print("#" * 70)

    # Build the object cache (loads all existing policy objects,
    # including those just uploaded in Phase 1)
    cache = ObjectCache(session, policy_object_profile_id)

    # Parse the NGFW Rules sheet from Excel
    # IMPORTANT: This returns a TUPLE of two dicts
    parse_result = parse_ngfw_rules_from_excel(EXCEL_FILE, cache)

    # ── Safety check: ensure we got a proper tuple of two dicts ──────────
    if parse_result is None:
        print("[PHASE 2] ❌ Parsing returned None. Exiting Phase 2.")
        return

    if isinstance(parse_result, tuple) and len(parse_result) == 2:
        zone_pair_rules, zone_pair_defaults = parse_result
    else:
        print(f"[PHASE 2] ❌ Unexpected return type from parser: {type(parse_result)}")
        print("[PHASE 2]    Expected tuple of (zone_pair_rules, zone_pair_defaults)")
        return

    # ── Validate that zone_pair_rules is a dict ──────────────────────────
    if not isinstance(zone_pair_rules, dict):
        print(f"[PHASE 2] ❌ zone_pair_rules is {type(zone_pair_rules)}, expected dict.")
        return

    if not isinstance(zone_pair_defaults, dict):
        print(f"[PHASE 2] ❌ zone_pair_defaults is {type(zone_pair_defaults)}, expected dict.")
        return

    if not zone_pair_rules:
        print("[PHASE 2] No zone pairs found in Excel. Exiting Phase 2.")
        return

    total_rules = sum(len(r) for r in zone_pair_rules.values())

    print(f"\n{'=' * 70}")
    print(f"  PHASE 2 PLAN: {MODE.upper()}")
    print(f"  Zone pairs:      {len(zone_pair_rules)}")
    print(f"  Total rules:     {total_rules}")
    print(f"  Default actions:")
    for zp in zone_pair_rules:
        da = zone_pair_defaults.get(zp, FALLBACK_DEFAULT_ACTION)
        print(f"    {zp[0]} -> {zp[1]}: {da.upper()}")
    if MODE == "update":
        print(f"  Profile:         {EXISTING_PROFILE_ID}")
        print(f"  Policy:          {EXISTING_POLICY_ID}")
    print(f"{'=' * 70}")

    confirm = input("\nProceed with Phase 2? [y/N]: ").strip().lower()
    if confirm != 'y':
        print("[PHASE 2] Cancelled by user.")
        return

    if MODE == "create":
        run_phase2_create(session, cache, zone_pair_rules, zone_pair_defaults)
    else:
        run_phase2_update(session, cache, zone_pair_rules, zone_pair_defaults)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    global VMANAGE_HOST, BASE_URL, USERNAME, PASSWORD
    global EXCEL_FILE, NEW_POLICY_NAME, NEW_POLICY_DESC
    global MODE, EXISTING_PROFILE_ID, EXISTING_POLICY_ID

    print("\n" + "=" * 70)
    print("  UNIFIED vMANAGE POLICY MANAGER")
    print("  Phase 1: Upload Policy Objects (Data Prefixes, Ports, Zones)")
    print("           Security Zones support Interface & VPN types")
    print("  Phase 2: Create/Update NGFW Policies (from 'NGFW Rules' sheet)")
    print("  Source:  Single Excel file (.xlsx)")
    print("=" * 70)

    # ── Collect Common Inputs ─────────────────────────────────────────────
    VMANAGE_HOST = input("\nEnter vManage Host (e.g., 198.168.1.10 or URL without https://): ").strip()
    while not VMANAGE_HOST:
        print("[ERROR] vManage Host cannot be empty.")
        VMANAGE_HOST = input("Enter vManage Host: ").strip()
    BASE_URL = f"https://{VMANAGE_HOST}"

    USERNAME = input("Enter Username: ").strip()
    while not USERNAME:
        print("[ERROR] Username cannot be empty.")
        USERNAME = input("Enter Username: ").strip()

    PASSWORD = getpass.getpass("Enter Password: ").strip()
    while not PASSWORD:
        print("[ERROR] Password cannot be empty.")
        PASSWORD = getpass.getpass("Enter Password: ").strip()

    # ── Single Excel File ────────────────────────────────────────────────
    EXCEL_FILE = input("Enter Excel File Name (.xlsx): ").strip()
    while not EXCEL_FILE:
        print("[ERROR] Excel File Name cannot be empty.")
        EXCEL_FILE = input("Enter Excel File Name (.xlsx): ").strip()

    if not os.path.isfile(EXCEL_FILE):
        print(f"\n[CRITICAL] Excel file '{EXCEL_FILE}' not found. Exiting.")
        sys.exit(1)

    # ── Phase Selection ──────────────────────────────────────────────────
    print("\n--- Phase Selection ---")
    print("  1 = Run Phase 1 only  (Upload policy objects)")
    print("  2 = Run Phase 2 only  (Create/Update NGFW policies)")
    print("  3 = Run Both Phases   (Phase 1 then Phase 2)")
    phase_choice = input("Select phases to run [3]: ").strip()
    if phase_choice not in ("1", "2", "3"):
        phase_choice = "3"

    run_p1 = phase_choice in ("1", "3")
    run_p2 = phase_choice in ("2", "3")

    # ── Phase 2 Inputs (only if Phase 2 will run) ────────────────────────
    if run_p2:
        print("\n--- Phase 2: NGFW Policy Configuration ---")
        NEW_POLICY_NAME = input("Enter New Policy Name: ").strip()
        while not NEW_POLICY_NAME:
            print("[ERROR] Policy Name cannot be empty.")
            NEW_POLICY_NAME = input("Enter New Policy Name: ").strip()

        NEW_POLICY_DESC = input("Enter New Policy Description: ").strip()
        if not NEW_POLICY_DESC:
            NEW_POLICY_DESC = f"NGFW Policy: {NEW_POLICY_NAME}"

        mode_input = input("Enter Mode ('create' or 'update') [create]: ").strip().lower()
        if mode_input in ("create", "update"):
            MODE = mode_input
        else:
            MODE = "create"

        if MODE == "update":
            EXISTING_PROFILE_ID = input("Enter Existing Profile ID: ").strip()
            while not EXISTING_PROFILE_ID:
                print("[ERROR] Profile ID is required for update mode.")
                EXISTING_PROFILE_ID = input("Enter Existing Profile ID: ").strip()

            EXISTING_POLICY_ID = input("Enter Existing Policy ID: ").strip()
            while not EXISTING_POLICY_ID:
                print("[ERROR] Policy ID is required for update mode.")
                EXISTING_POLICY_ID = input("Enter Existing Policy ID: ").strip()

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print(f"  vManage Host : {VMANAGE_HOST}")
    print(f"  Username     : {USERNAME}")
    print(f"  Password     : {'*' * len(PASSWORD)}")
    print(f"  Excel File   : {EXCEL_FILE}")
    print(f"  Run Phase 1  : {'Yes' if run_p1 else 'No'}")
    print(f"  Run Phase 2  : {'Yes' if run_p2 else 'No'}")
    if run_p2:
        print(f"  Policy Name  : {NEW_POLICY_NAME}")
        print(f"  Policy Desc  : {NEW_POLICY_DESC}")
        print(f"  Mode         : {MODE.upper()}")
        if MODE == "update":
            print(f"  Profile ID   : {EXISTING_PROFILE_ID}")
            print(f"  Policy ID    : {EXISTING_POLICY_ID}")
    print("=" * 70)

    # ── Verify required sheets exist ─────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        available_sheets = wb.sheetnames
        wb.close()
        print(f"\n  Available sheets in '{EXCEL_FILE}': {available_sheets}")

        if run_p1:
            for sn in ["Data Prefixes", "Port Lists"]:
                if sn in available_sheets:
                    print(f"    ✅ '{sn}' found")
                else:
                    print(f"    ⚠️ '{sn}' NOT found (will be skipped in Phase 1)")
            # Security Zones - note the 3-column format
            if "Security Zones" in available_sheets:
                print(f"    ✅ 'Security Zones' found (3-column: Name, Type, Values)")
            else:
                print(f"    ⚠️ 'Security Zones' NOT found (will be skipped in Phase 1)")

        if run_p2:
            if NGFW_RULES_SHEET in available_sheets:
                print(f"    ✅ '{NGFW_RULES_SHEET}' found")
            else:
                print(f"    ❌ '{NGFW_RULES_SHEET}' NOT found — Phase 2 will have no rules!")
    except Exception as e:
        print(f"\n  ⚠️ Could not preview sheets: {e}")

    confirm = input("\nProceed with the above settings? (yes/no): ").strip().lower()
    if confirm not in ("yes", "y"):
        print("Aborted by user.")
        sys.exit(0)

    # ══════════════════════════════════════════════════════════════════════
    # EXECUTION
    # ══════════════════════════════════════════════════════════════════════

    session = authenticate()

    policy_object_profile_id = get_policy_object_feature_profile_id(session)
    if not policy_object_profile_id:
        print("\n[CRITICAL] ❌ Could not retrieve Policy Object Feature Profile ID.")
        print("           Verify the policy-object profile exists in vManage.")
        sys.exit(1)

    if run_p1:
        run_phase1(session, policy_object_profile_id)
    else:
        print("\n[PHASE 1] ⏭️ Skipped by user.")

    if run_p2:
        run_phase2(session, policy_object_profile_id)
    else:
        print("\n[PHASE 2] ⏭️ Skipped by user.")

    print("\n" + "=" * 70)
    print("  ALL PHASES COMPLETE")
    print("=" * 70)


if __name__ == "__main__":
    main()
