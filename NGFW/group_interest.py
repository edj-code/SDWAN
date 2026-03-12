import requests
import json
import sys
import urllib3
import os
import openpyxl

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class Authentication:
    @staticmethod
    def get_jsessionid(vmanage_host, username, password):
        api = "/j_security_check"
        base_url = f"https://{vmanage_host}"
        url = base_url + api
        print(f"Authenticating to {url}...")
        payload = {'j_username': username, 'j_password': password}

        try:
            response = requests.post(url=url, data=payload, verify=False)
            response.raise_for_status()
            cookies = response.headers.get("Set-Cookie")
            if not cookies:
                print("Login failed. Check credentials.")
                sys.exit()
            jsessionid = cookies.split(";")[0]
            print("Login successful. JSESSIONID obtained.")
            return jsessionid
        except requests.exceptions.RequestException as e:
            print(f"Error during authentication: {e}")
            sys.exit()

    @staticmethod
    def get_token(vmanage_host, jsessionid):
        headers = {'Cookie': jsessionid}
        api = "/dataservice/client/token"
        base_url = f"https://{vmanage_host}"
        url = base_url + api

        response = requests.get(url=url, headers=headers, verify=False)
        if response.status_code == 200 and response.text:
            return response.text
        else:
            print(f"Failed to get token. Status: {response.status_code}")
            return None


def save_json_to_file(list_name, list_type, payload):
    """Save the JSON payload to a file for debugging/auditing purposes."""
    output_dir = "generated_policy_objects"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    filename = os.path.join(output_dir, f"{list_type}_{list_name}.json")
    try:
        with open(filename, 'w') as f:
            json.dump(payload, f, indent=4)
        print(f"   [FILE] Saved JSON to {filename}")
    except IOError as e:
        print(f"   [ERROR] Could not save JSON file: {e}")


# ============================================================
# STEP 1: GET THE EXISTING PARENT FEATURE PROFILE
# ============================================================
def get_policy_object_feature_profile_id(vmanage_host, jsessionid, token):
    """
    Finds the EXISTING global Feature Profile of type 'policy-object'.
    """
    base_url = f"https://{vmanage_host}"
    api_endpoint = "/dataservice/v1/feature-profile/sdwan/policy-object"
    url = base_url + api_endpoint
    headers = {
        'Cookie': jsessionid,
        'X-XSRF-TOKEN': token,
        'Content-Type': 'application/json'
    }

    print("\n[INIT] Searching for existing Feature Profile...")

    try:
        response = requests.get(url=url, headers=headers, verify=False)

        if response.status_code == 200:
            data = response.json()

            profiles = []
            if isinstance(data, dict):
                profiles = data.get('data', [])
            elif isinstance(data, list):
                profiles = data

            if len(profiles) > 0:
                existing_profile = profiles[0]
                p_id = existing_profile.get('profileId')
                p_name = existing_profile.get('profileName')
                print(f"[INIT] SUCCESS: Found existing global profile '{p_name}'. ID: {p_id}")
                return p_id
            else:
                print("[INIT] No profiles found in the list.")
                return None
        else:
            print(f"[ERROR] API call failed. Status: {response.status_code}")
            print(f"Response: {response.text}")
            return None

    except Exception as e:
        print(f"[ERROR] Exception during profile search: {e}")
        return None


# ============================================================
# STEP 2: READ DATA FROM EXCEL SHEETS
# ============================================================
def read_sheet_data(excel_file_path, sheet_name):
    """
    Reads a specific sheet from an Excel file.
    Returns a list of tuples: (list_name, [values]).
    Column 1 = Name of the list
    Column 2 = Values (single value or comma-separated)
    """
    data = []
    try:
        wb = openpyxl.load_workbook(excel_file_path, read_only=True)

        if sheet_name not in wb.sheetnames:
            print(f"   [WARNING] Sheet '{sheet_name}' not found in '{excel_file_path}'. Skipping.")
            wb.close()
            return data

        ws = wb[sheet_name]

        for row_num, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            # Skip empty rows
            if not row or row[0] is None:
                continue

            list_name = str(row[0]).strip()
            # Sanitize the list name (alphanumeric, underscores, hyphens)
            list_name = "".join(c for c in list_name if c.isalnum() or c in ('_', '-'))

            raw_values = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            values = [v.strip() for v in raw_values.split(',') if v.strip()]

            if list_name and values:
                data.append((list_name, values))

        wb.close()
        print(f"   [READ] Found {len(data)} entries in sheet '{sheet_name}'.")

    except FileNotFoundError:
        print(f"   [ERROR] File '{excel_file_path}' not found.")
    except Exception as e:
        print(f"   [ERROR] Failed to read sheet '{sheet_name}': {e}")

    return data


# ============================================================
# STEP 3: GENERIC API CALL TO CREATE A POLICY OBJECT PARCEL
# ============================================================
def create_policy_object_parcel(vmanage_host, jsessionid, token, parent_id, list_type, list_name, payload):
    """
    Generic function that POSTs a parcel to:
    /dataservice/v1/feature-profile/sdwan/policy-object/{parent_id}/{list_type}
    """
    base_url = f"https://{vmanage_host}"
    api_endpoint = f"/dataservice/v1/feature-profile/sdwan/policy-object/{parent_id}/{list_type}"
    url = base_url + api_endpoint

    headers = {
        'Content-Type': 'application/json',
        'Cookie': jsessionid,
        'X-XSRF-TOKEN': token
    }

    save_json_to_file(list_name, list_type, payload)

    try:
        print(f"   [API] Sending POST request for '{list_name}' (type: {list_type})...")
        response = requests.post(url=url, headers=headers, data=json.dumps(payload), verify=False)

        if response.status_code in [200, 201]:
            resp_json = response.json()
            obj_id = resp_json.get('id')
            print(f"   [SUCCESS] Created '{list_name}' (type: {list_type}). Parcel ID: {obj_id}")
        elif response.status_code == 400:
            print(f"   [FAIL] Error 400: Schema validation failed for '{list_name}'.")
            print(f"   Response: {response.text}")
        else:
            print(f"   [FAIL] Error {response.status_code} for '{list_name}'.")
            print(f"   Response: {response.text}")

    except requests.exceptions.RequestException as e:
        print(f"   [ERROR] Request error for '{list_name}': {e}")


# ============================================================
# STEP 4: BUILD PAYLOADS FOR EACH LIST TYPE
# ============================================================

def build_security_data_ip_prefix_payload(list_name, values):
    """
    Builds the JSON payload for 'security-data-ip-prefix' lists.
    Each value is an IP prefix (e.g., '10.0.0.0/12').
    """
    formatted_entries = []
    for prefix in values:
        entry = {
            "ipPrefix": {
                "optionType": "global",
                "value": prefix
            }
        }
        formatted_entries.append(entry)

    payload = {
        "name": list_name,
        "description": f"Security Data IP Prefix List: {list_name}",
        "data": {
            "entries": formatted_entries
        }
    }
    return payload


def build_security_port_payload(list_name, values):
    """
    Builds the JSON payload for 'security-port' lists.
    Each value is a port number (e.g., '80', '443').
    """
    formatted_entries = []
    for port in values:
        entry = {
            "port": {
                "optionType": "global",
                "value": port
            }
        }
        formatted_entries.append(entry)

    payload = {
        "name": list_name,
        "description": f"Security Port List: {list_name}",
        "data": {
            "entries": formatted_entries
        }
    }
    return payload


def build_security_zone_payload(list_name, values):
    """
    Builds the JSON payload for 'security-zone' lists.
    Each value is an interface identifier for the zone.
    """
    formatted_entries = []
    for interface in values:
        entry = {
            "interface": {
                "optionType": "global",
                "value": interface
            }
        }
        formatted_entries.append(entry)

    payload = {
        "name": list_name,
        "description": f"Security Zone: {list_name}",
        "data": {
            "entries": formatted_entries
        }
    }
    return payload


# ============================================================
# STEP 5: PROCESS EACH SHEET AND UPLOAD
# ============================================================

def process_sheet(vmanage_host, jsessionid, token, parent_id, excel_file, sheet_name, list_type, payload_builder):
    """
    Reads a sheet from the Excel file and creates policy object parcels
    for each row using the specified list_type and payload builder function.
    """
    print(f"\n{'='*60}")
    print(f"Processing sheet: '{sheet_name}' -> list type: '{list_type}'")
    print(f"{'='*60}")

    entries = read_sheet_data(excel_file, sheet_name)

    if not entries:
        print(f"   [INFO] No entries to process for sheet '{sheet_name}'.")
        return

    for idx, (list_name, values) in enumerate(entries, start=1):
        print(f"\n--- [{sheet_name}] Row {idx}: '{list_name}' with {len(values)} value(s) ---")
        payload = payload_builder(list_name, values)
        create_policy_object_parcel(
            vmanage_host, jsessionid, token, parent_id, list_type, list_name, payload
        )


def process_excel_and_upload(excel_file_path, vmanage_host, username, password):
    """
    Main orchestrator function:
    1. Authenticates to vManage
    2. Gets the parent Feature Profile ID
    3. Processes each sheet (Data Prefixes, Port Lists, Security Zones)
    """
    # 1. Authenticate
    jsessionid = Authentication.get_jsessionid(vmanage_host, username, password)
    token = Authentication.get_token(vmanage_host, jsessionid)

    if not token:
        print("[CRITICAL] Could not obtain XSRF token. Exiting.")
        return

    # 2. Get the Parent Container ID (Feature Profile)
    parent_profile_id = get_policy_object_feature_profile_id(vmanage_host, jsessionid, token)

    if not parent_profile_id:
        print("\n[CRITICAL] Could not obtain a valid Feature Profile ID. Exiting.")
        return

    # 3. Define the mapping of sheets to list types and payload builders
    sheet_config = [
        {
            "sheet_name": "Data Prefixes",
            "list_type": "security-data-ip-prefix",
            "payload_builder": build_security_data_ip_prefix_payload
        },
        {
            "sheet_name": "Port Lists",
            "list_type": "security-port",
            "payload_builder": build_security_port_payload
        },
        {
            "sheet_name": "Security Zones",
            "list_type": "security-zone",
            "payload_builder": build_security_zone_payload
        }
    ]

    # 4. Process each sheet
    for config in sheet_config:
        process_sheet(
            vmanage_host=vmanage_host,
            jsessionid=jsessionid,
            token=token,
            parent_id=parent_profile_id,
            excel_file=excel_file_path,
            sheet_name=config["sheet_name"],
            list_type=config["list_type"],
            payload_builder=config["payload_builder"]
        )

    print(f"\n{'='*60}")
    print("All sheets processed successfully.")
    print(f"{'='*60}")


# ============================================================
# MAIN ENTRY POINT
# ============================================================
if __name__ == "__main__":
    # Configuration
    VMANAGE_HOST = "10.89.1.130:44132"       # e.g., "198.18.1.10"
    USERNAME = "admin"           # e.g., "admin"
    PASSWORD = "cisco"           # e.g., "admin"
    EXCEL_FILE = "data.xlsx"         # e.g., "policy_objects.xlsx"

    process_excel_and_upload(EXCEL_FILE, VMANAGE_HOST, USERNAME, PASSWORD)